#!/usr/bin/env python3
"""Web Tasarım Freelance — Satış & Fiyatlandırma çalışma kitabı üreteci.

Tek kaynak: bu script. Çalıştır -> Satış Planı.xlsx hem vault'a hem OneDrive'a yazılır.
Her oturumda Akif teknik bilgi verdikçe ilgili sözlükler güncellenir, script yeniden çalışır.
"""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- yollar
VAULT_OUT = Path(__file__).resolve().parent / "Satış Planı.xlsx"
ONEDRIVE_OUT = Path.home() / "Library/CloudStorage/OneDrive-Kişisel/website proje/Satış Planı.xlsx"

# ---------------------------------------------------------------- stil
H1 = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
H2 = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD = Font(name="Calibri", size=11, bold=True)
NORMAL = Font(name="Calibri", size=11)
NOTE = Font(name="Calibri", size=10, italic=True, color="555555")

FILL_DARK = PatternFill("solid", fgColor="1F3B57")
FILL_MID = PatternFill("solid", fgColor="2E6DA4")
FILL_ACCENT = PatternFill("solid", fgColor="F2C14E")
FILL_LIGHT = PatternFill("solid", fgColor="EEF3F8")
FILL_INPUT = PatternFill("solid", fgColor="FFF7E0")

THIN = Side(style="thin", color="B7C4D0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def _title(ws, text, span):
    ws.merge_cells(f"A1:{get_column_letter(span)}1")
    c = ws["A1"]
    c.value = text
    c.font = H1
    c.fill = FILL_DARK
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26


def _header_row(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = H2
        c.fill = FILL_MID
        c.alignment = WRAP
        c.border = BOX


def _rows(ws, start, data, widths=None, accent_first=False):
    for r_off, rowvals in enumerate(data):
        row = start + r_off
        for i, val in enumerate(rowvals, start=1):
            c = ws.cell(row=row, column=i, value=val)
            c.font = BOLD if (accent_first and i == 1) else NORMAL
            c.alignment = WRAP
            c.border = BOX
            if r_off % 2 == 1:
                c.fill = FILL_LIGHT
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


# ================================================================ kitap
wb = Workbook()

# ---------------------------------------------------------------- 1. Özet / Strateji
ws = wb.active
ws.title = "Özet & Strateji"
_title(ws, "Web Tasarım Freelance — Satış Planı  |  Ekip: Akif (satış) + Dündar (teknik)", 4)

ws["A3"] = "TEMEL FİKİR: Ürünü önce üretme. Önce sat, kapora al, sonra üret."
ws["A3"].font = BOLD
ws.merge_cells("A3:D3")

_header_row(ws, 5, ["Aşama", "Kim", "Ne yapılır", "Çıktı / Hedef"])
_rows(ws, 6, [
    ["1. Hedefleme", "Akif", "Sektör seç (kuaför / kafe / klinik). Bölgede 20-30 aday esnaf listele.", "Aday listesi"],
    ["2. Hazır demo", "Dündar", "O sektöre 1 adet Basic şablon demo sitesi (canlı link).", "Gösterilebilir canlı demo"],
    ["3. Saha / arama", "Akif", "Esnafa git veya ara. Demoyu tablet/telefondan göster. İhtiyacı dinle.", "Randevu / ilgi"],
    ["4. Teklif", "Dündar", "Seçilen modülleri Teklif Hesaplayıcı'da topla, tek sayfa teklif ver.", "Fiyatlı teklif"],
    ["5. Kapora", "Akif", "%40-50 kapora al, işe o zaman başla. Kalan teslimde.", "Nakit giriş"],
    ["6. Üretim", "Dündar", "Şablonu müşteriye göre özelleştir, domain al, canlıya al, teslim.", "Yayında site"],
    ["7. Abonelik", "Dündar", "Bakım / reklam aboneliği sat. Tekrarlayan gelir.", "Aylık gelir"],
], widths=[16, 10, 52, 26], accent_first=True)

ws["A15"] = "ACİL HEDEF: Hafta 1-2'de 3-4 müşteri, 15-30k₺ nakit."
ws["A15"].font = BOLD
ws["A15"].fill = FILL_ACCENT
ws.merge_cells("A15:D15")

ws["A17"] = ("Neden 'üretmeden sat' çalışır: müşteri somut bir demo görüyor, hayal etmiyor. "
             "Kapora riski müşteriye taşıyor. Sen boşa şablon üretmiyorsun; her üretim bir "
             "satışın karşılığı. Tek sektöre 1 demo -> o sektördeki herkese aynı demo.")
ws["A17"].font = NOTE
ws["A17"].alignment = WRAP
ws.merge_cells("A17:D20")

# ---------------------------------------------------------------- 2. Fiyatlandırma
ws = wb.create_sheet("Fiyatlandırma")
_title(ws, "Modüler Fiyatlandırma — Müşteri ne isterse o kadarını öder", 4)

ws["A3"] = "1) BASE SITE (anahtar teslim, tek index sayfası — domain + canlıya alma dahil)"
ws["A3"].font = BOLD
_header_row(ws, 4, ["Paket", "Tasarım tipi", "Hazırlanma", "Fiyat (₺)"])
_rows(ws, 5, [
    ["Basic", "Standart tasarım, responsive, interaktif", "2-3 gün", 2000],
    ["Standart", "Gelişmiş tasarım, daha fazla özellik", "4-7 gün", 5000],
    ["Premium", "Video-scroll animasyonlu, profesyonel", "10-14 gün", 12000],
], widths=[18, 46, 14, 14], accent_first=True)

ws["A9"] = "2) SAYFA EKLEME (sayfa başı — Ürünler, Hakkımızda, İletişim vb.)"
ws["A9"].font = BOLD
_header_row(ws, 10, ["Paket", "Karmaşıklık", "Fiyat / sayfa (₺)", ""])
_rows(ws, 11, [
    ["Basic", "Standart sayfa (basit layout)", 500, ""],
    ["Standart", "Gelişmiş sayfa (filtreleme, carousel)", 1000, ""],
    ["Premium", "Kompleks sayfa (animasyon, interaktif)", 2000, ""],
], accent_first=True)

ws["A15"] = "3) AYLIK ABONELİKLER (bağımsız — müşteri hiçbirini ya da hepsini seçebilir)"
ws["A15"].font = BOLD
_header_row(ws, 16, ["Hizmet", "Basic (₺/ay)", "Standart (₺/ay)", "Premium (₺/ay)"])
_rows(ws, 17, [
    ["Bakım (içerik update + monitoring + QA)", 800, 1200, 1800],
    ["Premium Support 7/24 (danışmanlık + on-call)", 1500, 2500, 4000],
    ["Reklam Yönetimi (Google/Meta Ads + rapor)", 1500, 2500, 4000],
], accent_first=True)

ws["A22"] = "4) E-COMMERCE (online satış — backend + payment gateway)"
ws["A22"].font = BOLD
_header_row(ws, 23, ["Paket", "Özellik", "Anahtar teslim (₺)", "Aylık"])
_rows(ws, 24, [
    ["Basic", "Basit ürün sayfası + checkout", 3000, "500₺ + %2,5 komisyon"],
    ["Standart", "Kategoriler, sepet, stok takibi", 6000, "800₺ + %2 komisyon"],
    ["Premium", "Gelişmiş stok, müşteri hesabı, CRM", 10000, "1.200₺ + %1,5 komisyon"],
], accent_first=True)

# ---------------------------------------------------------------- 3. Teklif Hesaplayıcı
ws = wb.create_sheet("Teklif Hesaplayıcı")
_title(ws, "Teklif Hesaplayıcı — sarı hücreleri doldur, toplam otomatik", 4)

ws["A3"] = "Müşteri:"
ws["A3"].font = BOLD
ws["B3"].fill = FILL_INPUT
ws["B3"].border = BOX
ws["A4"] = "Tarih:"
ws["A4"].font = BOLD
ws["B4"].fill = FILL_INPUT
ws["B4"].border = BOX

_header_row(ws, 6, ["Kalem", "Birim fiyat (₺)", "Adet", "Tutar (₺)"])
calc = [
    ("Base Site — Basic", 2000),
    ("Base Site — Standart", 5000),
    ("Base Site — Premium", 12000),
    ("Ek sayfa — Basic", 500),
    ("Ek sayfa — Standart", 1000),
    ("Ek sayfa — Premium", 2000),
    ("E-Commerce — Basic", 3000),
    ("E-Commerce — Standart", 6000),
    ("E-Commerce — Premium", 10000),
]
for i, (name, price) in enumerate(calc):
    row = 7 + i
    ws.cell(row=row, column=1, value=name).border = BOX
    pc = ws.cell(row=row, column=2, value=price)
    pc.border = BOX
    pc.number_format = "#,##0"
    qc = ws.cell(row=row, column=3, value=0)
    qc.fill = FILL_INPUT
    qc.border = BOX
    qc.alignment = CENTER
    tc = ws.cell(row=row, column=4, value=f"=B{row}*C{row}")
    tc.border = BOX
    tc.number_format = "#,##0"

total_row = 7 + len(calc)
ws.cell(row=total_row, column=1, value="ANAHTAR TESLİM TOPLAM").font = BOLD
sum_c = ws.cell(row=total_row, column=4, value=f"=SUM(D7:D{total_row-1})")
sum_c.font = BOLD
sum_c.fill = FILL_ACCENT
sum_c.number_format = "#,##0"
sum_c.border = BOX
ws.cell(row=total_row, column=1).fill = FILL_ACCENT
ws.cell(row=total_row, column=1).border = BOX

kap_row = total_row + 1
ws.cell(row=kap_row, column=1, value="Kapora oranı (0-1)").font = BOLD
kr = ws.cell(row=kap_row, column=3, value=0.4)
kr.fill = FILL_INPUT
kr.border = BOX
kr.alignment = CENTER
kr.number_format = "0%"
ws.cell(row=kap_row, column=4, value=f"=D{total_row}*C{kap_row}").number_format = "#,##0"
ws.cell(row=kap_row, column=4).border = BOX
ws.cell(row=kap_row, column=1).border = BOX

ws.cell(row=kap_row + 1, column=1, value="Teslimde kalan").font = BOLD
ws.cell(row=kap_row + 1, column=4, value=f"=D{total_row}-D{kap_row}").number_format = "#,##0"
ws.cell(row=kap_row + 1, column=4).border = BOX
ws.cell(row=kap_row + 1, column=1).border = BOX

# aylık blok
m0 = kap_row + 3
ws.cell(row=m0, column=1, value="AYLIK ABONELİK (opsiyonel)").font = H2
ws.cell(row=m0, column=1).fill = FILL_MID
for col in range(2, 5):
    ws.cell(row=m0, column=col).fill = FILL_MID
_header_row(ws, m0 + 1, ["Hizmet", "Birim (₺/ay)", "Seç (1/0)", "Tutar (₺/ay)"])
monthly = [
    ("Bakım", 800),
    ("Premium Support 7/24", 1500),
    ("Reklam Yönetimi", 1500),
    ("E-Commerce aylık", 500),
]
for i, (name, price) in enumerate(monthly):
    row = m0 + 2 + i
    ws.cell(row=row, column=1, value=name).border = BOX
    pc = ws.cell(row=row, column=2, value=price)
    pc.border = BOX
    pc.number_format = "#,##0"
    sc = ws.cell(row=row, column=3, value=0)
    sc.fill = FILL_INPUT
    sc.border = BOX
    sc.alignment = CENTER
    tc = ws.cell(row=row, column=4, value=f"=B{row}*C{row}")
    tc.border = BOX
    tc.number_format = "#,##0"
mtot = m0 + 2 + len(monthly)
ws.cell(row=mtot, column=1, value="AYLIK TOPLAM").font = BOLD
ws.cell(row=mtot, column=1).fill = FILL_ACCENT
ws.cell(row=mtot, column=1).border = BOX
mc = ws.cell(row=mtot, column=4, value=f"=SUM(D{m0+2}:D{mtot-1})")
mc.font = BOLD
mc.fill = FILL_ACCENT
mc.number_format = "#,##0"
mc.border = BOX

for col, w in zip("ABCD", [30, 16, 12, 16]):
    ws.column_dimensions[col].width = w

# ---------------------------------------------------------------- 4. Müşteri Takip
ws = wb.create_sheet("Müşteri Takip")
_title(ws, "Müşteri Pipeline", 8)
_header_row(ws, 3, ["Müşteri", "Sektör", "İletişim", "Durum",
                    "Teklif (₺)", "Kapora alındı?", "Sonraki adım", "Not"])
ws.freeze_panes = "A4"
_rows(ws, 4, [["", "", "", "", "", "", "", ""] for _ in range(20)],
      widths=[22, 14, 18, 16, 12, 14, 26, 30])
ws["J3"] = "Durum seçenekleri: Aday / Görüşüldü / Teklif verildi / Kapora alındı / Üretimde / Teslim / Abonelik"
ws["J3"].font = NOTE

# ---------------------------------------------------------------- 5. Teknik Notlar
ws = wb.create_sheet("Teknik Notlar")
_title(ws, "Teknik Notlar — Akif'in verdiği bilgiler + ortak kararlar", 4)
_header_row(ws, 3, ["Konu", "Akif'in verdiği bilgi", "Dündar yorumu / karar", "Durum"])
_rows(ws, 4, [
    ["Hosting / domain", "", "", "Açık"],
    ["Şablon altyapısı (HTML mi, WP mi?)", "", "", "Açık"],
    ["Ödeme altyapısı (kapora nasıl alınacak)", "", "", "Açık"],
    ["Sözleşme / fatura", "", "", "Açık"],
    ["Reklam hesabı yönetimi", "", "", "Açık"],
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
], widths=[30, 40, 40, 12], accent_first=True)

# ---------------------------------------------------------------- yaz
wb.save(VAULT_OUT)
ONEDRIVE_OUT.parent.mkdir(parents=True, exist_ok=True)
shutil.copyfile(VAULT_OUT, ONEDRIVE_OUT)
print(f"yazıldı:\n  {VAULT_OUT}\n  {ONEDRIVE_OUT}")
